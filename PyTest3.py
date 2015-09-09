import urllib2, sys, string
from BeautifulSoup import BeautifulSoup
from collections import Counter
import re

#Pulling all text from webpage's html file
site= "http://liveramp.com"
hdr = {'User-Agent': 'Mozilla/5.0'}
req = urllib2.Request(site,headers=hdr)
page = urllib2.urlopen(req)
soup = BeautifulSoup(page)
words = soup.findAll(text=True)

#function to remove non-visible text from variable
def visible(w):
	if w.parent.name in ['style', 'script', '[document]', 'head', 'title']:
		return False
	elif re.match('<!--.*-->', str(w)):
		return False
	return True

#running function with original html text
visibleWords = filter(visible, words)

#removing commonly used words from the list
ignore = [' ', '\n', 'none', 'at', 'on', 'a' , 'the' , 'your', 'to', 'and', 'by', 'of', 'is', 'you', 'more', 'with', 'for']
wordCount = Counter(visibleWords)
for word in list(wordCount):
	if word in ignore:
		del wordCount[word]


#printing the 15 most commonly used words/phrases
topWords = wordCount.most_common(15)
print topWords


#putting data into excel document (Doesn't work yet)
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl import range
from openpyxl import get_column_letter


wb = Workbook()
wb.save('testExcel.xlsx')
ws = wb.active

r = 2
for stats in topWords:
	ws.cell (row=r, column=0).value = stats
	r +=1












